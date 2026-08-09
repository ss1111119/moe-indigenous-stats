"""科系差異與成長：原住民學生 vs 一般生，106–114 學年。

分子（原民）來自出版品：
  sheet A1-3 = 學科類別（27 個學門）
  sheet A1-4 = 細學類（約 150 類）＋其下的科系名稱明細
分母來自 opendata sdata（各校科系別概況），用科系代碼前 3／5 碼對上。
兩邊都是「中華民國學科標準分類（第 5 次修正）」。

「一般生」沿用教育部提要分析的定義：全體學生扣掉原住民學生。

輸出 out/：
  compare_field.csv / compare_major.csv   逐年逐等級的人數與占比
  growth_field.csv  / growth_major.csv    106→114 的成長與結構變化

先跑 `python fetch.py ebooks sdata`。
"""

import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).parent
DATA = ROOT / "data"
OUT = ROOT / "out"

# sdata 103–105 用 6 碼舊代碼（第 4 次修正），跟出版品的分類接不起來
FIRST_YEAR = 106
LAST_YEAR = 114

LEVELS = ["總計", "博士班", "碩士班", "學士班", "二專", "五專"]

LEVEL_FROM_CODE = {
    "D": "博士班",
    "M": "碩士班",
    "B": "學士班",
    "C": "學士班",
    "X": "學士班",
    "2": "二專",
    "5": "五專",
}

# 999 其他學門／99999 其他細學類：出版品含空大及進修學校、宗教研修學院，
# sdata 只有 140 所大專校院不含這些，沒有科系可歸類的學生全堆在這裡。
UNCOMPARABLE_PREFIX = "999"


def read_sheet(path: Path, sheet: str) -> list[list]:
    """.xls 走 xlrd、.xlsx 走 openpyxl，統一回傳純值的二維陣列。"""
    if path.suffix == ".xls":
        import xlrd

        ws = xlrd.open_workbook(path).sheet_by_name(sheet)
        return [[c.value for c in ws.row(i)] for i in range(ws.nrows)]

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    return [list(r) for r in wb[sheet].iter_rows(values_only=True)]


def _emit(year: int, code: str, name: str, nums, rows: list) -> None:
    for i, lv in enumerate(LEVELS):
        rows.append(
            {
                "學年度": str(year),
                "代碼": code,
                "名稱": name,
                "等級別": lv,
                "原民在學數": int(nums[i]),
                "原民畢業數": int(nums[i + 6]),
            }
        )


def parse_a13(path: Path, year: int) -> pd.DataFrame:
    """表 A1-3（學門）→ 長表。

    各年欄位位置不一致（106 的代碼是 '011' 且中間多一個空欄，114 是數字 11
    且沒有空欄），所以每列先濾掉空白格再按位置取值。
    """
    rows: list[dict] = []
    for raw in read_sheet(path, "A1-3"):
        cells = [c for c in raw if c not in (None, "")]
        if len(cells) < 13 or not isinstance(cells[1], str):
            continue
        code, name = cells[0], cells[1].strip()
        if "學門" not in name or not str(code).strip().isdigit():
            continue
        _emit(year, str(int(code)).zfill(3), name, cells[2:], rows)
    if not rows:
        raise ValueError(f"{path.name} 的 A1-3 解析不到任何學門列")
    return pd.DataFrame(rows)


def parse_a14(path: Path, year: int) -> pd.DataFrame:
    """表 A1-4（細學類）→ 長表。

    細學類是標題列（代碼在第 1 欄、名稱以「細學類」結尾），其下縮排的
    科系名稱明細列只有名稱沒有代碼，這裡只取標題列。
    """
    rows: list[dict] = []
    for raw in read_sheet(path, "A1-4"):
        cells = [c for c in raw if c not in (None, "")]
        if len(cells) < 13 or not isinstance(cells[1], str):
            continue
        code, name = str(cells[0]).strip(), cells[1].strip()
        if not name.endswith("細學類") or not code.isdigit():
            continue
        _emit(year, code.zfill(5), name, cells[2:], rows)
    if not rows:
        raise ValueError(f"{path.name} 的 A1-4 解析不到任何細學類列")
    return pd.DataFrame(rows)


def load_indigenous(parser) -> pd.DataFrame:
    frames = []
    for year in range(FIRST_YEAR, LAST_YEAR + 1):
        ext = "xls" if year <= 108 else "xlsx"
        path = DATA / "ebook" / f"{year}indigenous.{ext}"
        if not path.exists():
            print(f"  ! 缺 {path.name}，跳過")
            continue
        frames.append(parser(path, year))
    return pd.concat(frames, ignore_index=True)


def _sdata() -> pd.DataFrame:
    df = pd.read_json(DATA / "sdata.json", dtype=str, encoding="utf-8-sig")
    df["學年度"] = df["學年度"].astype(str)
    df = df[df["學年度"].astype(int).between(FIRST_YEAR, LAST_YEAR)].copy()
    df["等級別"] = df["等級別"].str[0].map(LEVEL_FROM_CODE)
    df = df[df["等級別"].notna()]
    for c in ("學生數", "上學年度畢業生數"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df


def load_all_students(digits: int) -> pd.DataFrame:
    """sdata → 學年度 × 分類 × 等級別的全體學生數（含「總計」列）。"""
    df = _sdata()
    df["代碼"] = df["科系代碼"].str[:digits]
    by_level = df.groupby(["學年度", "代碼", "等級別"], as_index=False)[
        ["學生數", "上學年度畢業生數"]
    ].sum()
    total = df.groupby(["學年度", "代碼"], as_index=False)[
        ["學生數", "上學年度畢業生數"]
    ].sum()
    total["等級別"] = "總計"
    out = pd.concat([by_level, total], ignore_index=True)
    return out.rename(columns={"學生數": "全體在學數", "上學年度畢業生數": "全體畢業數"})


def denominator_coverage() -> pd.Series:
    """sdata 每年的學生總數 ÷ 校別檔（student）同年總數。

    兩份檔案統計同一批人，正常應該完全相等；比值明顯小於 1 表示 sdata 那年有漏。
    """
    s = _sdata()
    a = pd.concat(
        [
            pd.read_json(DATA / f"{n}.json", dtype=str, encoding="utf-8-sig")
            for n in ("103-112_student", "student")
        ],
        ignore_index=True,
    )
    a["學年度"] = a["學年度"].astype(str)
    a["n"] = pd.to_numeric(a["總計"], errors="coerce").fillna(0)
    return (s.groupby("學年度")["學生數"].sum() / a.groupby("學年度")["n"].sum()).round(4)


def build(parser, digits: int) -> pd.DataFrame:
    ind = load_indigenous(parser)
    m = ind.merge(load_all_students(digits), on=["學年度", "代碼", "等級別"], how="left")

    # 一般生＝全體－原民，跟教育部提要分析的口徑一致。
    # 註：原民端含空大／宗教研修學院、全體端不含，相減會多扣約 0.14% 的全體人數。
    m["一般生在學數"] = m["全體在學數"] - m["原民在學數"]
    m["一般生畢業數"] = m["全體畢業數"] - m["原民畢業數"]
    m["原民佔比"] = (m["原民在學數"] / m["全體在學數"]).round(5)
    m["可比"] = ~m["代碼"].str.startswith(UNCOMPARABLE_PREFIX)

    # 結構占比：各自母體裡有多少比例念這個類別。分母排除不可比的 999，
    # 否則其他類別的占比會被那一坨堆積稀釋。
    ok = m[m["可比"]]
    tot = ok.groupby(["學年度", "等級別"])[["原民在學數", "一般生在學數"]].sum()
    idx = pd.MultiIndex.from_frame(m[["學年度", "等級別"]])
    for src, dst in (("原民在學數", "原民結構占比"), ("一般生在學數", "一般生結構占比")):
        m[dst] = (m[src].to_numpy() / tot[src].reindex(idx).to_numpy()).round(5)
    m.loc[~m["可比"], ["原民結構占比", "一般生結構占比"]] = pd.NA

    m["結構差距_百分點"] = ((m["原民結構占比"] - m["一般生結構占比"]) * 100).round(2)
    m["相對倍數"] = (m["原民結構占比"] / m["一般生結構占比"]).round(3)
    m["分母完整度"] = m["學年度"].map(denominator_coverage())
    return m.sort_values(["學年度", "等級別", "代碼"])


def build_growth(df: pd.DataFrame) -> pd.DataFrame:
    """106 → 114 的人數成長與結構變化（只看「總計」等級）。"""
    cur = df[(df["等級別"] == "總計") & df["可比"]]
    a = cur[cur["學年度"] == str(FIRST_YEAR)].set_index("代碼")
    b = cur[cur["學年度"] == str(LAST_YEAR)].set_index("代碼")
    keep = a.index.intersection(b.index)
    a, b = a.loc[keep], b.loc[keep]

    g = pd.DataFrame({"名稱": b["名稱"]})
    for who in ("原民", "一般生"):
        col = f"{who}在學數"
        g[f"{who}_{FIRST_YEAR}"] = a[col]
        g[f"{who}_{LAST_YEAR}"] = b[col]
        # 106 年為 0 的類別算不出成長率（細學類有幾個是新設的），留白
        base = a[col].astype(float).replace(0, float("nan"))
        g[f"{who}_成長率"] = (b[col].astype(float) / base - 1).round(4)
        s = f"{who}結構占比"
        g[f"{who}_占比變化_百分點"] = ((b[s] - a[s]) * 100).round(2)

    g["成長率差_百分點"] = ((g["原民_成長率"] - g["一般生_成長率"]) * 100).round(1)
    g[f"相對倍數_{FIRST_YEAR}"] = a["相對倍數"]
    g[f"相對倍數_{LAST_YEAR}"] = b["相對倍數"]
    g["相對倍數變化"] = (b["相對倍數"] - a["相對倍數"]).round(3)
    return g.reset_index().sort_values(f"{'原民'}_{LAST_YEAR}", ascending=False)


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    OUT.mkdir(exist_ok=True)

    for label, parser, digits, stem in (
        ("學門", parse_a13, 3, "field"),
        ("細學類", parse_a14, 5, "major"),
    ):
        df = build(parser, digits)
        df.to_csv(OUT / f"compare_{stem}.csv", index=False, encoding="utf-8-sig")
        growth = build_growth(df)
        growth.to_csv(OUT / f"growth_{stem}.csv", index=False, encoding="utf-8-sig")
        years = sorted(df["學年度"].unique())
        print(f"compare_{stem}.csv  {len(df):,} 列，{years[0]}–{years[-1]} 學年，"
              f"{df['代碼'].nunique()} 個{label}"
              f"　|　growth_{stem}.csv  {len(growth):,} 列")

    bad = denominator_coverage()
    for year, cov in bad[bad < 0.99].items():
        print(f"  ! {year} 學年 sdata 分母只有校別檔的 {cov:.1%}，該年占比偏高，勿採用")


if __name__ == "__main__":
    main()
