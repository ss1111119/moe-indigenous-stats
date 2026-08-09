"""把原住民學生數併上全體學生數，算出各校／全國的原住民學生佔比。

輸出到 out/：
  compare_by_school.csv   學年度 × 學校 × 等級別，原住民 vs 全體
  compare_national.csv    學年度 × 等級別 × 性別，全國彙總
  compare_grade.csv       學年度 × 等級別，各年級與延修生的人數結構（原民 vs 全體）

先跑 fetch.py 把 data/ 準備好。
"""

import re
import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).parent
DATA = ROOT / "data"
OUT = ROOT / "out"

# A1-1 的 5 個等級欄位 ←→ student.json 等級別代碼（取第一個字元）
# B=學士/四技、C=二技/二年制、X=4+X 都歸入學士班；5=五專與七年一貫歸入五專。
LEVEL_FROM_CODE = {
    "D": "博士班",
    "M": "碩士班",
    "B": "學士班",
    "C": "學士班",
    "X": "學士班",
    "2": "二專",
    "5": "五專",
}
LEVELS = ["博士班", "碩士班", "學士班", "二專", "五專"]

# A2-3 的等級別用詞逐年不一致，一併正規化到上面 5 個桶
A23_LEVEL = {
    "博士班": "博士班",
    "碩士班": "碩士班",
    "學士班": "學士班",
    "大學四年制": "學士班",
    "大學四年制(或四技)": "學士班",
    "大學二年制(或二技)": "學士班",
    "二技": "學士班",
    "4+X": "學士班",
    "二專": "二專",
    "五專": "五專",
}


# 學校代碼固定 4 碼，可能夾英文：0001 一般、0A01 附設進修專校、1R02 宗教研修學院
CODE_RE = re.compile(r"[0-9][0-9A-Za-z]{3}")


def isnum(x) -> bool:
    try:
        float(str(x).replace(",", ""))
        return True
    except ValueError:
        return False


def to_float(x) -> float:
    """出版品的數字有時是文字、有時是數值，缺值是「-」之類的符號。"""
    return float(str(x).replace(",", "")) if isnum(x) else 0.0


def read_sheet(path: Path, sheet: str) -> list:
    """.xls 走 xlrd、.xlsx 走 openpyxl；工作表名稱容許前後空白（113 學年是 'A1-1 '）。"""
    if path.suffix == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path)
        name = next(s for s in wb.sheet_names() if s.strip() == sheet)
        ws = wb.sheet_by_name(name)
        return [[c.value for c in ws.row(i)] for i in range(ws.nrows)]

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    name = next(s for s in wb.sheetnames if s.strip() == sheet)
    return [list(r) for r in wb[name].iter_rows(values_only=True)]


def to_num(s: pd.Series) -> pd.Series:
    """開放資料的數字是字串，缺值用 '-' 之類的符號表示。"""
    return pd.to_numeric(s.astype(str).str.replace(",", ""), errors="coerce").fillna(0)


def pad_code(s: pd.Series) -> pd.Series:
    """student.json 當年度檔的學校代碼沒補零（'1'），歷年檔與原民檔是 '0001'。"""
    return s.astype(str).str.strip().str.zfill(4)


def parse_ebook_a11(path: Path, year: int) -> pd.DataFrame:
    """出版品的表 A1-1（按等級別與校別分）→ 長表。

    opendata 的 A1-1 只更新到 113 學年，但出版品每年都有這張表，
    所以 114 學年的校別資料要從這裡拿。

    版面：學校類別是一列只有名稱＋數字的標題列，其下才是「代碼＋校名＋數字」的
    學校列，靠第 2 個非空欄位是不是字串來分辨。數字欄依序為
    在學總計 + 5 個等級，接著畢業總計 + 5 個等級。
    """
    raw = read_sheet(path, "A1-1")

    rows, category = [], None
    for r in raw:
        cells = [c for c in r if c not in (None, "")]
        if len(cells) < 7:
            continue
        head = str(cells[0]).strip()
        # 類別小計列：校名的位置是數字。學校列：校名是非數字的文字。
        if isnum(cells[1]):
            if head != "總計":
                category = head
            continue
        # 學校代碼是 4 碼，可能夾英文（0A01 附設進修專校、1R02 宗教研修學院）；
        # 沒有前導零的代碼會被存成數字，所以先轉字串再補零。
        code = head if not isnum(head) else str(int(float(head)))
        if not CODE_RE.fullmatch(code.zfill(4)) or len(cells) < 8:
            continue
        for i, lv in enumerate(LEVELS):            # cells[2] 是在學總計，跳過
            rows.append({
                "學年度": str(year), "學校類別": category,
                "學校代碼": code, "學校名稱": str(cells[1]).strip(), "等級別": lv,
                "原住民在學數": to_float(cells[3 + i]),
            })
    if not rows:
        raise ValueError(f"{path.name} 的 A1-1 解析不到任何學校列")
    return pd.DataFrame(rows)


def load_indigenous() -> pd.DataFrame:
    """A1-1 → 長表：學年度 / 學校代碼 / 學校名稱 / 等級別 / 原住民在學數。

    以 opendata 為主，出版品只用來補 opendata 還沒涵蓋的學年度；
    重疊的年份會拿來對帳，數字不合就出聲，避免版面解析錯了卻沒人發現。
    """
    df = pd.read_json(DATA / "indigenous_students_A1-1.json", dtype=str)
    cols = {f"在學學生人數_{lv}": lv for lv in LEVELS}
    for c in cols:
        df[c] = to_num(df[c])
    long = df.melt(
        id_vars=["學年度", "學校類別", "學校代碼", "學校名稱"],
        value_vars=list(cols),
        var_name="等級別",
        value_name="原住民在學數",
    )
    long["等級別"] = long["等級別"].map(cols)
    long["學校代碼"] = pad_code(long["學校代碼"])
    long["學年度"] = long["學年度"].astype(str)

    have = set(long["學年度"])
    extra = []
    for path in sorted((DATA / "ebook").glob("*indigenous.xls*")):
        year = int(path.name[:3])
        try:
            e = parse_ebook_a11(path, year)
        except Exception as exc:
            print(f"  ! {path.name} A1-1 解析失敗：{exc}")
            continue
        e["學校代碼"] = pad_code(e["學校代碼"])
        if str(year) in have:
            check(long, e, year)
        else:
            extra.append(e)
            print(f"  + {year} 學年校別資料取自出版品 A1-1（opendata 尚未更新）")
    return pd.concat([long, *extra], ignore_index=True) if extra else long


def check(opendata: pd.DataFrame, ebook: pd.DataFrame, year: int) -> None:
    """重疊年份對帳：出版品解析的結果應該和 opendata 完全相同。"""
    key = ["學校代碼", "等級別"]
    a = opendata[opendata["學年度"] == str(year)].set_index(key)["原住民在學數"]
    b = ebook.set_index(key)["原住民在學數"]
    common = a.index.intersection(b.index)
    diff = (a.loc[common] - b.loc[common]).abs()
    bad = int((diff > 0.5).sum())
    if bad or len(a) != len(b):
        print(f"  ! {year} 學年出版品與 opendata 不一致："
              f"{bad} 筆數字不同，列數 {len(a)} vs {len(b)}")


def load_all_students() -> pd.DataFrame:
    """student.json + 103-112_student.json → 長表：同鍵值的全體在學數。"""
    frames = []
    for name in ("103-112_student", "student"):
        d = pd.read_json(DATA / f"{name}.json", dtype=str, encoding="utf-8-sig")
        frames.append(d)
    df = pd.concat(frames, ignore_index=True)
    df["學年度"] = df["學年度"].astype(str)
    df["學校代碼"] = pad_code(df["學校代碼"])
    df["等級別"] = df["等級別"].str[0].map(LEVEL_FROM_CODE)
    df = df[df["等級別"].notna()]
    for c in ("總計", "男生計", "女生計"):
        df[c] = to_num(df[c])
    # A1-1 不分日間∕進修，這裡把各學制加總
    return df.groupby(
        ["學年度", "學校代碼", "學校名稱", "等級別"], as_index=False
    )[["總計", "男生計", "女生計"]].sum()


def build_by_school() -> pd.DataFrame:
    ind = load_indigenous()
    alls = load_all_students()
    m = ind.merge(
        alls.rename(columns={"總計": "全體在學數"})[
            ["學年度", "學校代碼", "等級別", "全體在學數"]
        ],
        on=["學年度", "學校代碼", "等級別"],
        how="left",
    )
    m["原住民佔比"] = (m["原住民在學數"] / m["全體在學數"]).round(5)
    return m.sort_values(["學年度", "學校代碼", "等級別"])


def build_national(by_school: pd.DataFrame) -> pd.DataFrame:
    """全國彙總（只用兩邊都對得上的列，避免分子有、分母缺造成失真）。"""
    ok = by_school[by_school["全體在學數"].notna()]
    nat = ok.groupby(["學年度", "等級別"], as_index=False)[
        ["原住民在學數", "全體在學數"]
    ].sum()
    nat["原住民佔比"] = (nat["原住民在學數"] / nat["全體在學數"]).round(5)
    # 109 學年起「○○科技大學附設進修專校」不再是獨立學校，其學生改列母校進修部，
    # 二專的分母因此一次跳增 2.5 倍。108 以前與 109 以後的二專佔比不可直接比較。
    nat["口徑註記"] = ""
    nat.loc[(nat["等級別"] == "二專") & (nat["學年度"] < "109"), "口徑註記"] = (
        "109前口徑：未含附設進修專校"
    )
    return nat


def build_grade() -> pd.DataFrame:
    """A2-3（原民，含 114 學年）對上全體的年級結構，看延修／各年級分布差異。"""
    grades = ["一年級", "二年級", "三年級", "四年級", "五年級", "六年級", "七年級"]
    ind = pd.read_json(DATA / "indigenous_students_A2-3.json", dtype=str)
    ind = ind[ind["學校類別"] == "大專校院"].copy()
    ind["等級別"] = ind["等級別"].map(A23_LEVEL)
    ind = ind[ind["等級別"].notna()]
    ind["學年度"] = ind["學年度"].astype(str)
    for g in grades:
        ind[g] = to_num(ind[f"{g}學生人數"])
    ind["延修生"] = to_num(ind["延修生學生人數"])
    ind_n = ind.groupby(["學年度", "等級別"], as_index=False)[grades + ["延修生"]].sum()
    ind_n = ind_n.melt(
        ["學年度", "等級別"], var_name="年級", value_name="原住民在學數"
    )

    frames = [
        pd.read_json(DATA / f"{n}.json", dtype=str, encoding="utf-8-sig")
        for n in ("103-112_student", "student")
    ]
    a = pd.concat(frames, ignore_index=True)
    a["學年度"] = a["學年度"].astype(str)
    a["等級別"] = a["等級別"].str[0].map(LEVEL_FROM_CODE)
    a = a[a["等級別"].notna()]
    for g in grades + ["延修生"]:
        a[g] = to_num(a[f"{g}男"]) + to_num(a[f"{g}女"])
    a_n = a.groupby(["學年度", "等級別"], as_index=False)[grades + ["延修生"]].sum()
    a_n = a_n.melt(["學年度", "等級別"], var_name="年級", value_name="全體在學數")

    m = ind_n.merge(a_n, on=["學年度", "等級別", "年級"], how="left")
    m["原住民佔比"] = (m["原住民在學數"] / m["全體在學數"]).round(5)
    return m[m["全體在學數"].fillna(0) > 0]


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    OUT.mkdir(exist_ok=True)

    by_school = build_by_school()
    by_school.to_csv(OUT / "compare_by_school.csv", index=False, encoding="utf-8-sig")

    nat = build_national(by_school)
    nat.to_csv(OUT / "compare_national.csv", index=False, encoding="utf-8-sig")

    grade = build_grade()
    grade.to_csv(OUT / "compare_grade.csv", index=False, encoding="utf-8-sig")

    # 對不到分母的絕大多數是該校根本沒開那個學制（原民檔每校都固定攤 5 個等級）；
    # 真正有人卻沒分母的，只有空大與宗教研修學院——它們不在「大專校院學生數」檔裡。
    gap = by_school[by_school["全體在學數"].isna() & (by_school["原住民在學數"] > 0)]
    print(f"compare_by_school.csv  {len(by_school):,} 列"
          f"（{len(gap):,} 列有原民生但缺分母：{'、'.join(sorted(gap['學校類別'].unique()))}）")
    print(f"compare_national.csv   {len(nat):,} 列")
    print(f"compare_grade.csv      {len(grade):,} 列")

    last = nat["學年度"].max()
    print(f"\n全國 {last} 學年度原住民學生佔比")
    print(nat[nat["學年度"] == last].to_string(index=False))

    # 同一等級內，延修生各佔自己母體多少——這條比佔比本身更能看出差異
    g = grade[grade["學年度"].astype(str) == str(grade["學年度"].max())]
    tot = g.groupby("等級別")[["原住民在學數", "全體在學數"]].sum()
    ext = g[g["年級"] == "延修生"].set_index("等級別")[["原住民在學數", "全體在學數"]]
    ratio = (ext / tot).dropna()
    ratio.columns = ["原住民延修生比率", "全體延修生比率"]
    print(f"\n延修生比率（{g['學年度'].iloc[0]} 學年度）")
    print(ratio.round(4).to_string())

    niu = by_school[by_school["學校名稱"].str.contains("宜蘭大學", na=False)]
    niu = niu[(niu["學年度"] == niu["學年度"].max()) & (niu["原住民在學數"] > 0)]
    if len(niu):
        print(f"\n國立宜蘭大學 {niu['學年度'].iloc[0]} 學年度")
        print(niu[["等級別", "原住民在學數", "全體在學數", "原住民佔比"]]
              .to_string(index=False))


if __name__ == "__main__":
    main()
