"""族籍別：各族原住民學生的等級別與性別結構，104–114 學年。

來源是出版品的表 A1-2（按等級別、族籍別與性別分），沒有 opendata 版本。

這一塊**沒有一般生對照**——一般生沒有族籍。所以它不是「原民 vs 一般生」的比較，
而是回答另一個問題：把「原住民」當成單一群體看，掩蓋了多少差異。

版面：每一族占三列（計／男／女），族名放在中間那列，欄位位置逐年不同，
所以用「計／男／女」這個標記的位置反推，不寫死欄號。

輸出 out/ethnicity.csv。
"""

import sys
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).parent
DATA = ROOT / "data"
OUT = ROOT / "out"

LEVELS = ["總計", "博士班", "碩士班", "學士班", "二專", "五專"]
MARKS = ("計", "男", "女")


def read_sheet(path: Path, sheet: str) -> list:
    """.xls 走 xlrd、.xlsx 走 openpyxl；工作表名稱容許前後空白。"""
    if path.suffix == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name(next(s for s in wb.sheet_names() if s.strip() == sheet))
        return [[c.value for c in ws.row(i)] for i in range(ws.nrows)]

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    name = next(s for s in wb.sheetnames if s.strip() == sheet)
    return [list(r) for r in wb[name].iter_rows(values_only=True)]


def isnum(x) -> bool:
    try:
        float(str(x).replace(",", ""))
        return True
    except ValueError:
        return False


def parse_a12(path: Path, year: int) -> pd.DataFrame:
    rows, group = [], []
    for r in read_sheet(path, "A1-2"):
        cells = [(i, c) for i, c in enumerate(r) if c not in (None, "")]
        mark = next((i for i, c in cells
                     if isinstance(c, str) and c.strip() in MARKS), None)
        if mark is None:
            continue
        nums = [c for i, c in cells if i > mark and isnum(c)]
        if len(nums) < 12:
            continue
        label = next((str(c).replace("　", "").strip() for i, c in cells
                      if i < mark and isinstance(c, str) and c.strip()), None)
        sex = next(c.strip() for i, c in cells if i == mark)
        group.append({"sex": sex, "label": label, "nums": nums})

        if sex != "女":                       # 每族三列，女是最後一列
            continue
        name = next((g["label"] for g in group if g["label"]), None)
        if name:
            for g in group:
                if g["sex"] == "計":
                    continue                  # 計＝男＋女，不重複收
                for i, lv in enumerate(LEVELS):
                    rows.append({
                        "學年度": year, "族籍別": name, "性別": g["sex"],
                        "等級別": lv, "在學數": int(float(str(g["nums"][i]).replace(",", ""))),
                    })
        group = []

    if not rows:
        raise ValueError(f"{path.name} 的 A1-2 解析不到任何族籍列")
    return pd.DataFrame(rows)


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    OUT.mkdir(exist_ok=True)

    frames = []
    for path in sorted((DATA / "ebook").glob("*indigenous.xls*")):
        year = int(path.name[:3])
        try:
            frames.append(parse_a12(path, year))
        except Exception as exc:
            print(f"  ! {path.name}：{exc}")
    df = pd.concat(frames, ignore_index=True)

    # 對帳：各族加總應等於出版品自己的「總計」列
    chk = df[df["族籍別"] != "總計"].groupby(["學年度", "等級別"])["在學數"].sum()
    tot = df[df["族籍別"] == "總計"].groupby(["學年度", "等級別"])["在學數"].sum()
    gap = (chk - tot).abs()
    bad = gap[gap > 0]
    print(f"對帳：各族加總 vs 出版品總計，{len(bad)} 組不符" +
          ("" if bad.empty else f"\n{bad.to_string()}"))

    df.to_csv(OUT / "ethnicity.csv", index=False, encoding="utf-8-sig")
    years = sorted(df["學年度"].unique())
    names = [n for n in df["族籍別"].unique() if n != "總計"]
    print(f"ethnicity.csv  {len(df):,} 列，{years[0]}–{years[-1]} 學年，{len(names)} 族")

    last = years[-1]
    cur = (df[(df["學年度"] == last) & (df["等級別"] == "總計") & (df["族籍別"] != "總計")]
           .pivot_table(index="族籍別", columns="性別", values="在學數", aggfunc="sum")
           .fillna(0).astype(int))
    cur["合計"] = cur["男"] + cur["女"]
    cur["女性占比"] = (cur["女"] / cur["合計"] * 100).round(1)
    print(f"\n{last} 學年 各族大專在學人數")
    print(cur.sort_values("合計", ascending=False).to_string())


if __name__ == "__main__":
    main()
